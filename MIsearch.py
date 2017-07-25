import os
import sys
import timeit
import openpyxl

# main function
def main():
    # start program timer
    start_time = timeit.default_timer()

    # parses command line for files
    if (len(sys.argv) <= 1):
        print "ERROR: Please enter a filename."
        sys.exit()

    # iterates through all files
    for file_input in sys.argv[1:]:
        print "Analyzing: \"" + file_input + "\""

        # opens input file
        try:
            wb_input = openpyxl.load_workbook(file_input, data_only=True)
        except:
            print "ERROR: Invalid filename \"" + file_input + "\""
            continue

        # reads from input file
        sheet_input = wb_input.active

        # iterates through all rows of input spreadsheet
        interval1 = []  # array of NAT intervals (theoretical, charge, low, high, and ID)
        interval2 = []  # array of SIL intervals (theoretical, charge, low, high, and ID)
        data = []       # array of data (experimental value and charge)
        for row in sheet_input:
            # check if it is the first row
            if (row[0].value != "Experimental"):
                # appends the experimental value and charge (column A and B)
                data.append([row[0].value, row[1].value])
                # check if there is a valid interval
                if (row[2].value != None):
                    # appends the NAT interval theoretical, charge, low, high, and ID (columns C to G)
                    interval1.append([row[2].value, row[3].value, row[4].value, row[5].value, row[6].value])
                    # appends the SIL interval theoretical, charge, low, high, and ID (columns H to K, and G)
                    interval2.append([row[7].value, row[8].value, row[9].value, row[10].value, row[6].value])

        interval1.sort(key = lambda x: (x[1], x[2]))    # sorts the NAT intervals by charge and low value
        interval2.sort(key = lambda x: (x[1], x[2]))    # sorts the SIL intervals by charge and low value
        data.sort(key = lambda x: x[0])                 # sorts the data by experimental value

        # analyzes file
        output1 = []            # array of NAT output data
        output2 = []            # array of SIL output data
        max_index = len(data)   # length of experimental values

        # iterates through each NAT interval
        index = 0       # index of experimental value
        charge = 0      # default charge value
        for interval in interval1:
            # check if interval charge has changed
            if (interval[1] != charge):
                index = 0               # reset index to zero
                charge = interval[1]    # set the charge to the current charge
            # loop while index doesn't exceed the number of experimental values
            while (index < max_index):
                # check if data charge equals the interval charge AND
                # check if experimental value is between the low and high interval
                if (data[index][1] == interval[1] and data[index][0] >= interval[2] and data[index][0] <= interval[3]):
                    # append theoretical, charge, and ID to output data
                    output1.append([interval[0], interval[1], interval[4]])
                    # break and move to the next interval
                    break
                # check if experimental value is above the high interval
                if (data[index][0] > interval[3]):
                    # break and move to the next interval
                    break
                # increment index
                index += 1

        # iterates through each SIL interval
        index = 0       # index of experimental value
        charge = 0      # default charge value
        for interval in interval2:
            # check if interval charge has changed
            if (interval[1] != charge):
                index = 0               # reset index to zero
                charge = interval[1]    # set the charge to the current charge
            # loop while index doesn't exceed the number of experimental values
            while (index < max_index):
                # check if data charge equals the interval charge AND
                # check if experimental value is between the low and high interval
                if (data[index][1] == interval[1] and data[index][0] >= interval[2] and data[index][0] <= interval[3]):
                    # append theoretical, charge, and ID to output data
                    output2.append([interval[0], interval[1], interval[4]])
                    # break and move to the next interval
                    break
                # check if experimental value is above the high interval
                if (data[index][0] > interval[3]):
                    # break and move to the next interval
                    break
                # increment index
                index += 1

        output1.sort(key = lambda x: (x[1], x[2]))  # sorts the NAT output by charge and ID
        output2.sort(key = lambda x: (x[1], x[2]))  # sorts the SIL output by charge and ID

        # iterates through each NAT and SIL intervals, comparing them
        output = []     # combined NAT output (charge, NAT theoretical, SIL theoretical, and ID)
        for test1 in output1:
            for test2 in output2:
                # check if NAT and SIL IDs are the same AND
                # check if NAT and SIL charge are the same
                if (test1[2] == test2[2] and test1[1] == test2[1]):
                    # append charge, NAT theoretical, SIL theoretical, and ID to output data
                    output.append([test1[1], test1[0], test2[0], test2[2]])

        # writes to output file
        wb_output = openpyxl.Workbook()
        sheet_output = wb_output.active

        # writes the headers for the first row
        row_count = 1
        sheet_output.cell(row = row_count, column = 1).value = "Output-Charge"
        sheet_output.cell(row = row_count, column = 2).value = "Output-NAT Theoretical"
        sheet_output.cell(row = row_count, column = 3).value = "Output-SIL Theoretical"
        sheet_output.cell(row = row_count, column = 4).value = "Output-ID"

        # writes the charge, NAT theoretical, SIL theoretical, and ID for each row
        for value in output:
            row_count += 1
            sheet_output.cell(row = row_count, column = 1).value = value[0]
            sheet_output.cell(row = row_count, column = 2).value = value[1]
            sheet_output.cell(row = row_count, column = 3).value = value[2]
            sheet_output.cell(row = row_count, column = 4).value = value[3]

        # saves output file
        file_output = os.path.splitext(file_input)[0] + "_results.xlsx"

        try:
            print "Writing: \"" + file_output + "\""
            wb_output.save(file_output)
        except:
            print "ERROR: Please close output file."
            continue

    # end program timer
    end_time = timeit.default_timer()
    elapsed_time = end_time - start_time
    print "Elapsed time: {:0.2f} seconds".format(elapsed_time)

if __name__ == '__main__':
    main()
